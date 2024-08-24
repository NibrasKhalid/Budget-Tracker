#Nibras Khalid

"""Budget Tracker Script

This script is a simple budget tracking application designed to help manage
 finances by recording income and expenses. It utilizes Excel files to store 
and manage budget data. The script performs the following tasks:

The script provides a menu-driven interface to interact with the user, allowing them to
choose from adding new entries, viewing summaries, adding recurring transactions, generating
reports, or exiting the application. The script also validates inputs to ensure they conform
to expected formats and values."""


import re, time, subprocess
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import pyinputplus as pyip

# Define file paths for budget and report Excel files
budgetTracker = 'budgetTracker.xlsx'
reportsTracker = 'monthlyAnnualReports.xlsx'

# Function to create a new Excel file if it doesn't exist
# This function initializes a new Excel workbook with a single worksheet named "Budget",
# and sets up the header row for storing budget entries.
def createExcel():
    # Create a new workbook object
    wb = Workbook()
    # Select the active worksheet
    ws = wb.active
    # Rename the active worksheet to "Budget"
    ws.title = "Budget"
    # Append the header row for budget entries
    ws.append(["Date", "Category", "Description", "Amount"])
    # Save the workbook with the defined file name
    wb.save(budgetTracker)

# Function to read existing data from the Excel file
# This function attempts to load the Excel file containing budget data.
# If the file does not exist, it calls createExcel() to create a new one.
# It then reads and returns the data from the active worksheet.
def readBudgetTracker():
    try:
        # Attempt to load the existing workbook
        wb = load_workbook(budgetTracker)
    except FileNotFoundError:
        # If file not found, create a new Excel file
        createExcel()
        # Load the newly created workbook
        wb = load_workbook(budgetTracker)  
    # Select the active worksheet
    ws = wb.active
    # Extract all data from the worksheet into a list
    data = [row for row in ws.values]
    # Close the workbook to free resources
    wb.close()
    return data

# Function to write updated data to the Excel file
# This function writes the given data to the Excel file, creating a new workbook
# and saving the updated entries in the "Budget" worksheet. It then opens the file 
# for viewing and waits until it is closed.
def writeBudgetTracker(data):
    # Create a new workbook object
    wb = Workbook()
    # Select the active worksheet
    ws = wb.active
    # Rename the active worksheet to "Budget"
    ws.title = "Budget"
    # Append each row of data to the worksheet
    for row in data:
        ws.append(row)
    # Save the workbook with the defined file name
    wb.save(budgetTracker)
    
    # Open the Excel file and wait for it to close
    trackerWait = subprocess.Popen(['start', budgetTracker], shell=True)
    trackerWait.wait()

# Function to validate amount
# This function ensures that the amount entered is a positive number with up to two decimal places.
# It uses a regular expression to check the format and raises a ValueError if the format is invalid 
# or if the amount is less than or equal to zero.
def amountValidation(amount):
    # Convert amount to string for validation
    amountInput = str(amount)
    
    # Define regular expression for valid amount format
    regExp = r"^\d+(\.\d{1,2}?)$"
    
    # Check if amount matches the regular expression
    if not re.match(regExp, amountInput):
        raise ValueError("Invalid amount format. Please enter a positive number with up to two decimal places.")
    
    # Convert the amount string to float
    amountFloat = float(amountInput)
    # Ensure the amount is greater than zero
    if amountFloat <= 0:
        raise ValueError("Amount must be greater than zero.")
    
    return amountFloat

# Function to validate date format
# This function checks if the provided date string is in the correct format (YYYY-MM-DD).
# It returns True if the date is valid and False otherwise.
def dateValidation(date):
    try:
        # Attempt to parse the date string
        datetime.strptime(date, "%Y-%m-%d")
        return True
    except ValueError:
        # Return False if date parsing fails
        return False

# Function to add an entry
# This function adds a new entry to the budget. It validates the amount and appends the new
# entry (date, category, description, amount) to the existing budget data, then writes
# the updated data back to the Excel file.
def addEntry(date, category, description, amount):
    # Validate the amount
    amount = amountValidation(amount)
    # Create a new entry with the provided data
    newEntry = [date, category, description, amount]
    # Read existing budget data
    data = readBudgetTracker()
    # Append the new entry to the data
    data.append(newEntry)
    # Write the updated data back to the Excel file
    writeBudgetTracker(data)  

# Function to generate monthly and annual reports
# This function calculates and creates summaries of income and expenses by month and year.
# It writes the results to a new Excel file, with separate sections for monthly and yearly summaries.
def generate_reports():
    # Read existing budget data
    data = readBudgetTracker()
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly and Annual Reports"
    
    # Initialize dictionaries to store monthly and yearly summaries
    monthlySummary = {}
    yearlySummary = {}

    # Iterate through each row of data (skipping the header row)
    for row in data[1:]:
        date, category, description, addAmount = row
        try:
            addAmount = float(addAmount)
            # Extract month and year from the date
            monthYear = datetime.strptime(date, "%Y-%m-%d").strftime("%Y-%m")
            year = datetime.strptime(date, "%Y-%m-%d").strftime("%Y")

            # Initialize monthly and yearly summary entries if not present
            if monthYear not in monthlySummary:
                monthlySummary[monthYear] = {'Income': 0, 'Expense': 0}
            if year not in yearlySummary:
                yearlySummary[year] = {'Income': 0, 'Expense': 0}

            # Update the appropriate summary based on the category
            if category.lower() == 'income':
                monthlySummary[monthYear]['Income'] += addAmount
                yearlySummary[year]['Income'] += addAmount
            elif category.lower() == 'expense':
                monthlySummary[monthYear]['Expense'] += addAmount
                yearlySummary[year]['Expense'] += addAmount
        except ValueError:
            # Skip rows with invalid amounts
            pass

    # Add headers and data for monthly summary
    ws.append(["Month", "Income", "Expense"])
    for monthYear, summary in monthlySummary.items():
        ws.append([monthYear, summary['Income'], summary['Expense']])

    # Add a blank row to separate sections
    ws.append([])
    # Add headers and data for yearly summary
    ws.append(["Year", "Income", "Expense"])
    for year, summary in yearlySummary.items():
        ws.append([year, summary['Income'], summary['Expense']])

    # Save the workbook with the defined file name
    wb.save(reportsTracker)
    
    # Open the report file and wait for it to close
    reportWait = subprocess.Popen(['start', reportsTracker], shell=True)   
    reportWait.wait()

# Function to calculate summary
# This function calculates the total income, total expenses, and balance by iterating
# through the budget data. It also computes the totals for each category (income and expense).
def summaryCalc():
    # Read existing budget data
    data = readBudgetTracker()
    incomeTotal = 0
    expenseTotal = 0
    categoryTotal = {"income": 0, "expense": 0}

    # Iterate through each row of data (skipping the header row)
    for row in data[1:]:
        date, category, description, amount = row
        # Convert amount to float
        amount = float(amount)
        # Update totals based on the category
        if category.lower() == 'income':
            incomeTotal += amount
            categoryTotal["income"] += amount
        elif category.lower() == 'expense':
            expenseTotal += amount
            categoryTotal["expense"] += amount

    # Calculate the remaining balance
    balance = incomeTotal - expenseTotal
    return incomeTotal, expenseTotal, balance, categoryTotal

# Function to add recurring transactions
# This function prompts the user to input details for a recurring transaction and adds 
# it to the budget for a specified number of occurrences, incrementing the date by 30 days 
# for each subsequent entry.
def recurringTransactions():
    # Prompt user for the start date of the recurring transactions
    date = input("Enter the date for the first occurrence (YYYY-MM-DD): ")
    while not dateValidation(date):
        # Re-prompt if the date format is invalid
        print("Invalid date format. Please enter again (YYYY-MM-DD).")
        date = input("Enter the date for the first occurrence (YYYY-MM-DD): ")

    # Prompt user for the category of the transaction
    category = input("Enter the category (Income/Expense): ")
    while category.lower() not in ["income", "expense"]:
        # Re-prompt if the category is invalid
        print("Invalid category. Please enter 'Income' or 'Expense'.")
        category = input("Enter the category (Income/Expense): ")

    # Prompt user for a description of the transaction
    description = input("Enter a description: ")

    # Prompt user for the amount of the transaction
    amount = input("Enter the amount: ")
    recurring = 1
    while recurring == 1:
        try:
            # Validate the amount input
            amount = amountValidation(amount)  
            recurring = 0
            break
        except ValueError as ve:
            # Re-prompt if the amount is invalid
            print(ve)
            amount = input("Enter the amount: ")

    # Prompt user for the number of recurrences
    countRecc = pyip.inputInt("Enter the number of recurrences: ", min=1)

    # Add the recurring transaction for each occurrence
    for i in range(countRecc):
        addEntry(date, category, description, amount)
        # Increment the date by 30 days for the next occurrence
        date = (datetime.strptime(date, "%Y-%m-%d") + timedelta(days=30)).strftime("%Y-%m-%d")  

    print(f"Added {countRecc} recurring transactions.")

# Main function
# This function controls the program's flow by displaying a menu and handling user
# choices. It provides options to add entries, show a summary report, add recurring
# transactions, generate reports, show budget tracker excel, or exit the application.
def mainFunc():
    try:
        with open(budgetTracker, 'r'):
            pass
    except FileNotFoundError:
        createExcel()
    
    # Initialize the user choice
    choice = None  
    while choice != '6':  
        # Display the menu options
        time.sleep(2)
        print("\n1. Add a new entry")
        print("2. Show summary report")
        print("3. Add recurring transactions")
        print("4. Generate monthly and annual reports")
        print("5. Open Budget Tracker Excel")
        print("6. Exit \n")
        
        # Prompt user for their choice
        choice = input("Enter your choice (1-6): ")   
        
        if choice == '1':
            # Handle option to add a new entry
            date = input("Enter the date (YYYY-MM-DD): ")
            while not dateValidation(date):
                print("Invalid date format. Please enter again (YYYY-MM-DD).")
                date = input("Enter the date (YYYY-MM-DD): ")

            category = input("Enter the category (Income/Expense): ")
            while category.lower() not in ["income", "expense"]:
                print("Invalid category. Please enter 'Income' or 'Expense'.")
                category = input("Enter the category (Income/Expense): ")

            description = input("Enter a description: ")

            amount = input("Enter the amount: ")
            while True:
                try:
                    amount = amountValidation(amount) 
                    break
                except ValueError as ve:
                    print(ve)
                    amount = input("Enter the amount: ")

            time.sleep(1)
            # Add the new entry to the budget
            addEntry(date, category, description, amount)
            print("\n\nEntry added successfully")

            time.sleep(1)  

        elif choice == '2':
            # Handle option to show a summary report
            incomeTotal, expenseTotal, balance, categoryTotal = summaryCalc()
            print(f"\n\nSummary Report:\nTotal Income: ${incomeTotal:.2f}\nTotal Expenses: ${expenseTotal:.2f}\nRemaining Balance: ${balance:.2f}")
            time.sleep(1)  

        elif choice == '3':
            # Handle option to add recurring transactions
            recurringTransactions()
            time.sleep(1)  

        elif choice == '4':
            # Handle option to generate monthly and annual reports
            generate_reports()
            time.sleep(1)  

        elif choice == '5':
            # Handle option to open the Budget Tracker Excel file
            budgetWait = subprocess.Popen(['start', budgetTracker], shell=True)
            budgetWait.wait()
            time.sleep(1)  

        elif choice == '6':  
            # Handle option to exit the program
            time.sleep(1)  
            print("\nExiting the program.\n")
            time.sleep(1)  
            break

        else:
            # Handle invalid menu choices
            print("Invalid choice. Please enter a number between 1 and 5.") 

print("Budget Tracker")
mainFunc()
